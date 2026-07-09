import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DASHBOARD_STAFF = [
    ["S1101","Md. Anwar","PC","1773","Male","anwar1773@gmail.com","8639399057","","Active","","CCTNS Req/Modifications; CCTNS-NCRB/TCS; CCTNS Reports; AP Smart Cops; APCOPS; BSNL DLT; AP Website; Trainings","Krishna District","","2024-01-01"],
    ["S1102","L. Trinadha Rao","PC","1722","Male","trinadh_pcs@appolice.gov.in","9652881207","","Active","","CCTNS BSNL Connectivity; iRAD Reports; XDR+SASE; GeM Bid Evaluations; Dispatch; District Coordination","Tirupati District","","2024-01-01"],
    ["S1103","A. Ganeswara Rao","PC","1705","Male","arnipalli17051@gmail.com","9704347908","","Active","","CCTNS & Pragathi Reports; AP Website/Suraksha; XDR+SASE; GeM Bid Evaluations; Dispatch; District Coordination","Tirupati District","","2024-01-01"],
    ["S1104","B. Siva Srinivas Rao","PC","5122","Male","sssvas05@gmail.com","7386104098","","Active","","CCTNS Application Statistics; AP Smart Cops; AI4AP; GeM Bid Evaluations; Dispatch; District Coordination","Guntur","","2024-01-01"],
    ["S1105","D. Aashik","PC","3582","Male","aashiqkhaleja1@gmail.com","8639592237","","Active","","CEIR; GeM Bid Evaluations; District Coordination; Trainings","NTR District","","2024-01-01"],
    ["S1106","M. Suneetha","WPC","2546","Female","suneethacctns@gmail.com","9948255139","","Active","","PM Gati Shakti Portal; BSNL DLT & eSMS; GeM Bid Evaluations; Drafting; Trainings; eOffice Follow-up","Krishna District","","2024-01-01"],
    ["S1107","PVK. Prasad","HC(C)","1649","Male","iampvkprasad@gmail.com","9110300554","","Active","","NATGRID - Gandiva Portal; E-Forms Approvals; GeM Bid Evaluations; Drafting; Trainings; District Coordination","Communication","","2024-01-01"],
    ["S1108","N. Bhanu Hari","PC(C)","1871","Male","bhanuhari.123@gmail.com","8309578086","","Active","","NATGRID - Sudarshan Portal; E-Forms Approvals; GeM Bid Evaluations; Drafting; Trainings; District Coordination","Communication","","2024-01-01"],
    ["S1109","Sk. Lal John Basha","HC(C)","1806","Male","laljanibasha99@gmail.com","9573180827","","Active","","NATGRID Gandiva & Sudarshan; Case Approvals/MHA Coordination; GeM Bid Evaluations; Trainings","Communication","","2024-01-01"],
    ["S1110","E. Sai Sowmya","WPC(C)","1884","Female","sowmyacctns@gmail.com","8317554950","","Active","","CEIR; GeM Bid Evaluations; Drafting; Office Inward/Outward; Meetings Coordination","Communication","","2024-01-01"],
    ["S1111","L. Naveen Babu","PC","2950","Male","naveenbabu.lingam@gmail.com","8712345058","","Active","","CEIR; GeM Bid Evaluations; Drafting; VC/Training Video Uploads; District Coordination; Trainings","Krishna District","","2024-01-01"],
    ["S1112","S. Rajasekhar Rao","PC(C)","1817","Male","rajasekhararao27@gmail.com","9849583594","","Active","","CCTNS Helpdesk 24/7; JTrac Bug Tracking; DG DSR; Deletions/Modifications via DSP Login","Communication","","2024-01-01"],
    ["S1113","K. Gopi","PC","4300","Male","gopikavuris43@gmail.com","9553312439","","Active","","CCTNS Helpdesk 24/7 Monitoring","Guntur District","","2024-01-01"],
    ["S1114","L. Rahul","PC","2885","Male","rahullanka2@gmail.com","7095233770","","Active","","CCTNS Helpdesk 24/7 Monitoring","Krishna District","","2024-01-01"],
    ["S1115","K. Sreenu","PC","1644","Male","sreenu.kellakarrothu@gmail.com","6301615193","","Active","","CCTNS Helpdesk 24/7 Monitoring","Tirupathi","","2024-01-01"],
    ["S1116","N. Ramanjaneyulu","PC","2401","Male","nayuduramu332@gmail.com","9603557975","","Active","","CriMAC; APCOPS 2.0; eSakshya & eSummons; GeM Bid Evaluations; NCRB Coordination","Prakasam District","","2024-01-01"],
    ["S1117","J. Mounika","WPC","3786","Female","mounikajonnalagadda@gmail.com","6281029399","","Active","","eSakshya & eSummons; GeM Bid Evaluations; NCRB Coordination; District Coordination","NTR District","","2024-01-01"],
    ["S1118","E. Vasantha Kumari","WPC","3762","Female","vasu92028@gmail.com","8730267046","","Active","","eSakshya & eSummons; GeM Bid Evaluations; NCRB Coordination; District Coordination","NTR District","","2024-01-01"],
]

STAFF_TEMPLATE = [
    ["S1001","K. Vijaya Kumar","SI","","Male","","9440470479","2018-11-23","Active","","APOLIS, Drones, Mee-seva","Guntur Range","","2024-01-01"],
    ["S1002","V.Sheshanjali","PC","","Female","","","","Active","","APOLIS","Guntur Dist","","2024-01-01"],
    ["S1003","CH.Srinivasa Rao","HC","316","Male","","6281861526","2020-01-15","Active","","Drones, APFRS-Biometric, Meeseva","Guntur Dist","","2024-01-01"],
    ["S1004","M.Mohan","SI","","Male","","","","Active","","Data center","","","2024-01-01"],
    ["S1005","K. Sreekanth","RI","","Male","","9491416025","2021-04-04","Active","","Cloud based CCTVs, CCTV in PS, E-Office","APSP","","2024-01-01"],
    ["S1006","T.Koteswararao","PC","1815","Male","","9000640658","2019-01-23","Active","","CCTV in PS","Communication","","2024-01-01"],
    ["S1007","O.Jyothi","WPC","2067","Female","","7674033487","2019-01-23","Active","","Cloud based CCTVs","Communication","","2024-01-01"],
    ["S1008","CH.Venkatesh","PC","741","Male","","8886885859","2022-09-01","Active","","E-Office","Eluru Dist","","2024-01-01"],
    ["S1009","S.Srinivasulu","HC","1610","Male","","7013677239","2024-09-05","Active","","E-Office","Vizianagaram Dist","","2024-01-01"],
    ["S1010","G.Ravi kiran","SI","","Male","","","","Active","","E-challans, ITMS, ERSS-112, Tabs","","","2024-01-01"],
    ["S1011","R.Ramarao","PC","1615","Male","","9441473964","2022-07-26","Active","","ERSS-112 AND TABS, 48 call takers","Tirupathi Dist","","2024-01-01"],
    ["S1012","K.Kiran Vani","WHC","1982","Female","","8977959518","2019-01-23","Active","","E-challans, Brerath Analyzers, Speed Laser guns","Communication","","2024-01-01"],
    ["S1013","P.Ashok Kumar","PC","444","Male","","9160998444","2022-03-26","Active","","E-challans","Palnadu Dist","","2024-01-01"],
    ["S1014","T. Venkata Naidu","RSI","","Male","","6304218129","2022-10-12","Active","","BWC, Women safety (Disha), LHMS","Kadapa Dist","","2024-01-01"],
    ["S1015","CH.Siva Balakrishna","HC","1639","Male","","9491800714","2023-12-17","Active","","BWC, LHMS","Communication","","2024-01-01"],
    ["S1016","R.Kishore","PC","1435","Male","","9666881292","2023-10-01","Active","","Women Safety","6th Battalion","","2024-01-01"],
    ["S1017","E Prasanna","WPC","6872","Female","","9505328763","2025-04-09","Active","","Women Safety","Guntur Dist","","2024-01-01"],
    ["S1018","V.Sudharshana reddy","IC","","Male","","","","Active","","CCTNS","","","2024-01-01"],
    ["S1019","T. Anoj Kumar","RSI","","Male","","9493946020","2024-01-11","Active","","CCTNS, NATGRID, CEIR, BISAG","DAR Srikakulam","","2024-01-01"],
    ["S1020","Ram Koti Naik","ARSI","","Male","","9951929562","2018-03-15","Active","","MTO, 112 Control Room Shift","SCR CPL Amberpet","","2024-01-01"],
    ["S1021","G. Govinda Rao","PC","1146","Male","","9391211030","2024-12-15","Active","","MTO Writer","Tirupati District","","2024-01-01"],
    ["S1022","D. Raviner Babu","HC","917","Male","","8688887707","2022-06-14","Active","","Driver","PTO","","2024-01-01"],
    ["S1023","SK. Silar Vali","HC","997","Male","","9908069956","2019-06-14","Active","","Driver","PTO","","2024-01-01"],
    ["S1024","SK. Riyaz","PC","554","Male","","9493515387","","Active","","Driver","Nellore","","2024-01-01"],
    ["S1025","S. Ravi Kumar","PC","387","Male","","6303289717","2019-01-14","Active","","Driver - Director FSL","16th Bn. APSP VSKP","","2024-01-01"],
    ["S1026","G. Paramesh","PC","915","Male","","7013574236","2023-04-27","Active","","Driver","Palnadu District","","2024-01-01"],
    ["S1027","A. Syambabu","HG","202","Male","","9010634392","2024-04-10","Active","","112 Control Room (Driver)","Krishna District","","2024-01-01"],
    ["S1028","B. Rani","WASI","","Female","","9247836443","2022-10-07","Active","","112 Control Room Shift","West Godavari District","","2024-01-01"],
    ["S1029","Pallapothu Suresh","PC","6075","Male","","","","Active","","SAKTHI Team","Palnadu","","2024-01-01"],
    ["S1030","Ch.Durga Rao","PC","2669","Male","","","","Active","","SAKTHI Team","Eluru","","2024-01-01"],
    ["S1031","P Harika","WPC","2058","Female","","","","Active","","SAKTHI Team","Communication","","2024-01-01"],
    ["S1032","P.Komala","WPC","2051","Female","","","","Active","","SAKTHI Team","Communication","","2024-01-01"],
    ["S1033","A.Prasanna Lakshmi","WPC","2075","Female","","","","Active","","SAKTHI Team","Communication","","2024-01-01"],
    ["S1034","T.Keerthi","WPC","2057","Female","","","","Active","","SAKTHI Team","Communication","","2024-01-01"],
    ["S1035","Sarojini","AAO","","Female","","","","Active","","AO Section and Staff","","","2024-01-01"],
    ["S1036","L.Venkata Krishna","PC","4159","Male","","","","Active","","CCTNS Reports","Bapatla District","","2024-01-01"],
    ["S1037","Md.Sadhik","RSI","","Male","","","","Active","","Data Center-2","NTR Dist","","2024-01-01"],
    ["S1038","M. Gowri Shankara Rao","PC","198","Male","","","","Active","","Data center","Krishna Dist","","2024-01-01"],
    ["S1039","A. Murali","PC","956","Male","","","","Active","","Data center","16th BN","","2024-01-01"],
    ["S1040","S. Ramu","PC","1891","Male","","","","Active","","Data center","PCO","","2024-01-01"],
    ["S1041","Y. Pardha Saradhi","PC","4344","Male","","","","Active","","Data center","NTR Dist","","2024-01-01"],
    ["S1042","K. Ananth Kumar","PC","1305","Male","","","","Active","","Data center","6th BN","","2024-01-01"],
    ["S1043","P. Sunil Kumar","HC","1826","Male","","","","Active","","DGP Office Team","PCO","","2024-01-01"],
    ["S1044","K. Praveen","PC","8303","Male","","","","Active","","DGP Office Team","PCO","","2024-01-01"],
    ["S1045","K.Sathish","IC","","Male","","9248092990","2021-06-20","Active","","ICJS - eSakshya, eSummons, NDSO, CRIMAC, ITSSO, RTGS","NTR District","","2024-01-01"],
]

DUTIES_MAP = {
    "8639399057": {"d": "1. Handle CCTNS application requirements and bug reports. 2. Coordinate with NCRB & TCS team. 3. Prepare CCTNS reports. 4. Monitor BSNL DLT & eSMS.", "w": "1. Weekly Teleconference Report. 2. Update CCTNS statistics. 3. GeM bid evaluation.", "m": "1. Monthly CCTNS performance report. 2. Update project master sheet."},
    "9652881207": {"d": "1. Check email dig_pcs@appolice.gov.in. 2. Monitor CCTNS BSNL connectivity. 3. Handle dispatch and outward correspondence.", "w": "1. Weekly Teleconference Report. 2. GeM bid evaluation. 3. Submit iRAD reports.", "m": "1. Update XDR+SASE project status. 2. Review GeM procurement progress."},
    "9704347908": {"d": "1. Check emails dsp_cctns@ and cctns_ap@. 2. Prepare CCTNS & Pragathi reports. 3. Test Citizen Portal. 4. Handle dispatch.", "w": "1. Weekly VC Report on CCTNS and Citizen Portal. 2. GeM bid evaluation. 3. Update AP Police Website data.", "m": "1. Monthly Pragathi Reports. 2. Update XDR+SASE project details."},
    "7386104098": {"d": "1. Check email apsmartcops@appolice.gov.in. 2. Compile CCTNS statistics. 3. Prepare News360AI clippings. 4. Submit News360AI DSR. 5. Coordinate with PROs.", "w": "1. CCTNS Training at APSP 6th BN. 2. Smart Policing report. 3. GeM bid evaluation.", "m": "1. Coordinate district data. 2. Submit data with Sparity team. 3. AI4AP file correspondence."},
    "8639592237": {"d": "1. Transfer CEIR requests to other states. 2. Review CEIR statistics. 3. Monitor CEIR NIC email. 4. Review CEIR WhatsApp group. 5. Update CEIR Excel.", "w": "1. Smart Policing report on CEIR. 2. Communicate to low-performing units. 3. GeM evaluation. 4. Weekly VC Report.", "m": "1. Comparative CEIR assessment with other states. 2. Update CEIR master sheet."},
    "9948255139": {"d": "1. Maintain eOffice system. 2. Handle cctns_ap@ncrb.gov.in. 3. Monitor PM GatiShakti Portal. 4. Check pmgatishakti@appolice.gov.in. 5. Update PM GatiShakti Excel.", "w": "1. Monitor PM GatiShakti for GIS layer issues. 2. Monitor BSNL DLT Portal. 3. Weekly VC Report. 4. GeM evaluation.", "m": "1. Update Google Sheets tracker for PM GatiShakti and BSNL DLT."},
    "9110300554": {"d": "1. Approve cases in GANDIVA portal. 2. Examine VPN requests. 3. Onboard new users. 4. Resolve user issues. 5. Submit daily NATGRID status report.", "w": "1. Compile GANDIVA usage data. 2. Monitor inactive cases. 3. Weekly VC Report. 4. GeM evaluation.", "m": "1. Update Gandiva master sheet. 2. Collect success stories. 3. Conduct training sessions."},
    "8309578086": {"d": "1. Check email success_stories_pcs@. 2. Address NATGRID Sudarshan issues. 3. Resolve Sudarshan issues. 4. Update Sudarshan Excel.", "w": "1. Coordinate with least-performing units. 2. Weekly VC Report. 3. GeM evaluation.", "m": "1. Update Sudarshan master sheet. 2. Prepare inactive users list. 3. Check VPN renewals."},
    "9573180827": {"d": "1. Approve cases in GANDIVA. 2. Onboard new users. 3. Resolve user issues with NATGRID Delhi. 4. Submit daily NATGRID status report. 5. Check natgrid@appolice.gov.in.", "w": "1. Generate Gandiva usage report. 2. Review low-performing units. 3. Weekly VC Report. 4. GeM evaluation.", "m": "1. Update Gandiva master sheet. 2. Conduct training sessions."},
    "8317554950": {"d": "1. Check cctns_ap@ncrb.gov.in. 2. Update Meetings Sheet. 3. Process office inward/outward. 4. Update CEIR Excel.", "w": "1. Weekly VC Report. 2. GeM evaluation. 3. Review pending correspondence.", "m": "1. Monthly CEIR status update. 2. Compile inward/outward summary."},
    "8712345058": {"d": "1. Maintain eOffice systems. 2. Handle cctns_ap@ncrb.gov.in and technical_services@appolice.gov.in. 3. Process CEIR requests. 4. Upload VC/Training videos daily. 5. Update CEIR Excel.", "w": "1. Weekly VC Report. 2. GeM evaluation. 3. Dispatch and drafting. 4. Conduct trainings.", "m": "1. Monthly CEIR transfer status. 2. Review pending correspondence."},
    "9849583594": {"d": "1. Raise CCTNS bugs in JTrac. 2. Handle deletions/modifications via DSP login. 3. Submit Helpdesk DSR. 4. Update DG DSR Google Sheet. 5. Check helpdesk_pcs@appolice.gov.in.", "w": "1. Generate JTrac weekly report. 2. Follow up on open bugs with TCS team.", "m": "1. Monthly JTrac report. 2. Monthly Helpdesk performance summary."},
    "9553312439": {"d": "1. Monitor CCTNS 24/7 during shift. 2. Log and escalate incidents. 3. Verify CCTNS portal uptime.", "w": "1. Handover shift briefing. 2. Update helpdesk register.", "m": "1. Submit shift-wise performance note."},
    "7095233770": {"d": "1. Monitor CCTNS 24/7 during shift. 2. Log and escalate incidents. 3. Verify CCTNS portal uptime.", "w": "1. Handover shift briefing. 2. Update helpdesk register.", "m": "1. Submit shift-wise performance note."},
    "6301615193": {"d": "1. Monitor CCTNS 24/7 during shift. 2. Log and escalate incidents. 3. Verify CCTNS portal uptime.", "w": "1. Handover shift briefing. 2. Update helpdesk register.", "m": "1. Submit shift-wise performance note."},
    "9603557975": {"d": "1. Monitor Zero FIR Report in CriMAC. 2. Tech support for CriMAC issues. 3. Support districts in APCOPS 2.0. 4. Check apcops_pcs@ email.", "w": "1. Coordinate with low-performing units. 2. Weekly CriMAC performance report. 3. Weekly VC Report. 4. GeM evaluation.", "m": "1. Monthly Zero FIR transfer list. 2. Update APCOPS 2.0 master sheet."},
    "6281029399": {"d": "1. Coordinate with districts on eSakshya and eSummons issues. 2. Liaise with NCRB for updates.", "w": "1. Weekly VC Report. 2. GeM evaluation. 3. Review portal performance.", "m": "1. Update eSakshya master sheet. 2. Compile NCRB coordination summary."},
    "8730267046": {"d": "1. Coordinate with districts on eSakshya and eSummons issues. 2. Liaise with NCRB for updates.", "w": "1. Weekly VC Report. 2. GeM evaluation. 3. Review eSummons portal performance.", "m": "1. Update eSummons master sheet. 2. Compile NCRB coordination summary."},
}

# Merge (deduplicate by phone and name)
dash_phones = set(r[6].strip() for r in DASHBOARD_STAFF if r[6].strip())
dash_names = set(r[1].strip().lower() for r in DASHBOARD_STAFF)

all_staff = []
sno = 1
for r in DASHBOARD_STAFF:
    duties = DUTIES_MAP.get(r[6], {})
    all_staff.append([sno, 'Dashboard Sheet', r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[10], r[11], duties.get('d',''), duties.get('w',''), duties.get('m','')])
    sno += 1

for r in STAFF_TEMPLATE:
    ph = r[6].strip()
    nm = r[1].strip().lower()
    if ph not in dash_phones and nm not in dash_names:
        all_staff.append([sno, 'Duty Chart', r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[10], r[11], '', '', ''])
        sno += 1

print(f'Total staff: {len(all_staff)}')

# Create Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'All Staff'

# Header style
hdr_font = Font(bold=True, color='FFFFFF', size=11)
hdr_fill_d = PatternFill('solid', fgColor='0D3B7A')  # navy for dashboard
hdr_fill = PatternFill('solid', fgColor='0D3B7A')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ['S.No', 'Source', 'Staff ID', 'Full Name', 'Rank', 'Badge No', 'Gender', 'Gmail', 'Phone', 'Date Joined', 'Status', 'Assigned Projects', 'District / Unit', 'Daily Duties', 'Weekly Duties', 'Monthly Duties']
col_widths = [6, 14, 10, 22, 10, 10, 8, 30, 13, 14, 8, 45, 20, 60, 50, 45]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = border
    ws.column_dimensions[cell.column_letter].width = w

ws.row_dimensions[1].height = 30

# Data rows
fill_dash = PatternFill('solid', fgColor='EEF4FF')  # light blue for dashboard staff
fill_tmpl = PatternFill('solid', fgColor='FFFFFF')

for row_idx, row in enumerate(all_staff, 2):
    is_dash = row[1] == 'Dashboard Sheet'
    for col, val in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = fill_dash if is_dash else fill_tmpl
        cell.border = border
        cell.alignment = wrap if col >= 12 else center
    ws.row_dimensions[row_idx].height = 40 if row[13] else 20

# Freeze header
ws.freeze_panes = 'A2'

# Sheet 2: Dashboard staff with full duties
ws2 = wb.create_sheet('18 Core Staff (Full Details)')
headers2 = ['S.No', 'Staff ID', 'Full Name', 'Rank', 'Badge No', 'Gender', 'Gmail', 'Phone', 'District / Unit', 'Assigned Projects', 'Daily Duties', 'Weekly Duties', 'Monthly Duties']
col_widths2 = [6, 10, 22, 10, 10, 8, 30, 13, 20, 45, 70, 55, 50]

for col, (h, w) in enumerate(zip(headers2, col_widths2), 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = PatternFill('solid', fgColor='7B0D0D')
    cell.alignment = center
    cell.border = border
    ws2.column_dimensions[cell.column_letter].width = w
ws2.row_dimensions[1].height = 30

for i, r in enumerate(DASHBOARD_STAFF, 1):
    duties = DUTIES_MAP.get(r[6], {})
    row_data = [i, r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[11], r[10], duties.get('d',''), duties.get('w',''), duties.get('m','')]
    for col, val in enumerate(row_data, 1):
        cell = ws2.cell(row=i+1, column=col, value=val)
        cell.fill = PatternFill('solid', fgColor='FFF0F0') if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        cell.border = border
        cell.alignment = wrap if col >= 10 else center
    ws2.row_dimensions[i+1].height = 80

ws2.freeze_panes = 'A2'

out = r'C:\Users\prasa\Desktop\TS-Staff-System\TS_Staff_All_63_Members.xlsx'
wb.save(out)
print(f'Saved: {out}')
